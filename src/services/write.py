import os
from typing import Tuple

from PySide6.QtCore import QThread, Signal
from csv import DictReader, __version__ as csv_version
from openpyxl import load_workbook, __version__ as openpyxl_version
from pymupdf import open as open_media, Point, __version__ as pymupdf_version  # type: ignore

from utils.logger import logger

MAX_TEXT_LENGTH = 200


class WriterThread(QThread):
    outputSignal = Signal(str, str)

    def __init__(
        self,
        coordinateX: float,
        coordinateY: float,
        fontSize: float,
        rotation: int,
        pageNumber: int,
        color: Tuple[float, float, float],
        fontPath: str,
        documentPath: str,
        tablePath: str,
        outputPath: str,
    ):
        super().__init__()
        self.documentPath = documentPath
        self.tablePath = tablePath
        self.outputPath = outputPath
        self.fontPath = fontPath
        self.coordinateX = coordinateX
        self.coordinateY = coordinateY
        self.fontSize = fontSize
        self.color = color
        self.rotation = rotation
        self.pageNumber = pageNumber

    def output(self, text: str, level: str = "INFO"):
        logger.log(level, text)
        self.outputSignal.emit(text, level)

    def readTable(self) -> tuple[list[dict[str, str]], list[str]]:
        if self.tablePath.endswith(".csv"):
            logger.info(
                f"Using native CSV {csv_version} module to read {self.tablePath}"
            )
            records = []
            with open(self.tablePath, mode="r", encoding="utf-8") as csvfile:
                reader = DictReader(csvfile)
                headers = reader.fieldnames
                if not headers:
                    raise ValueError("CSV file has no headers")
                logger.info(f"Headers found in CSV: {len(headers)}")
                for row in reader:
                    records.append(
                        {
                            key: (value if value else "")
                            for key, value in row.items()
                        }
                    )
        elif self.tablePath.endswith(".xlsx"):
            logger.info(
                f"Using openpyxl {openpyxl_version} to read {self.tablePath}"
            )
            workbook = load_workbook(filename=self.tablePath, data_only=True)
            logger.info(
                f"Loaded workbook with {len(workbook.sheetnames)} sheets"
            )
            sheet = workbook.active
            if not sheet:
                raise ValueError("Excel file has no sheets")
            logger.info(
                f"Active sheet: {sheet.title}, {sheet.max_row} rows, {sheet.max_column} columns"
            )
            headers = [str(cell.value) for cell in sheet[1] if cell.value]
            logger.info(f"Headers found: {len(headers)}")
            if not headers:
                raise ValueError("Excel file has no headers")
            records = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # type: ignore
                logger.debug(f"Row data: {row}")
                record = {
                    headers[i]: (value if value else "")
                    for i, value in enumerate(row)
                    if i < len(headers)
                }
                records.append(record)
        else:
            raise ValueError("Unsupported file extension")

        return records, list(headers)

    def readDocument(self) -> bytes:
        try:
            logger.info(
                f"Using PyMuPDF {pymupdf_version} to read {self.documentPath}"
            )
            file = open_media(self.documentPath)
            data = file.convert_to_pdf()
            return data
        except Exception as e:
            self.output(f"Failed to read file: {e}", "ERROR")
            raise

    def run(self):
        inputs = self.__dict__.copy()
        logger.info(f"Starting writer thread with input parameters: {inputs}")
        self.output("...")

        with logger.catch():
            if not os.path.isdir(self.outputPath):
                self.output(
                    f"Output directory {self.outputPath} not found or is not a directory",
                    "ERROR",
                )
                return

            try:
                data = self.readDocument()
                logger.info(
                    f'Document loaded, extension is "{self.documentPath.split(".")[-1]}"'
                )
            except Exception as e:
                self.output(f"Failed to open file: {e}", "ERROR")
                return

            try:
                rows, cols = self.readTable()
            except Exception as e:
                self.output(f"Failed to load table: {e}", "ERROR")
                return

            logger.info(
                f"Table loaded, found {len(rows)} {len(rows) == 1 and 'row' or 'rows'}"
            )
            logger.info(f"Table columns read: {cols}")
            logger.info(f"Table rows read: {rows}")

            if not rows or not cols:
                self.output("Given table is empty", "ERROR")
                return
            if len(cols) != 1:
                self.output(
                    "Given table must have exactly one column with the text to write on each row",
                    "ERROR",
                )
                return
            if len(rows) < 1:
                self.output(
                    "Given table must have at least another row in addition to the headers!",
                    "ERROR",
                )
                return

            total: int = 0
            successful: int = 0
            for row in rows:
                total += 1
                text = row[cols[0]].strip()

                if not text:
                    self.output(
                        f"[{total}] Text to write is empty on this row", "ERROR"
                    )
                    continue
                if len(text) > MAX_TEXT_LENGTH:
                    self.output(
                        f"[{total}] Text to write has more than {MAX_TEXT_LENGTH} characters on this row",
                        "ERROR",
                    )
                    continue

                pdf = open_media("pdf", data)

                try:
                    fontName = os.path.basename(self.fontPath)
                    page = pdf[self.pageNumber]
                    page.clean_contents()
                    page.insert_font(fontfile=self.fontPath, fontname=fontName)
                    page.insert_text(  # type: ignore
                        Point(self.coordinateX, self.coordinateY),
                        text,
                        fontname=fontName,
                        fontsize=self.fontSize,
                        color=self.color,
                        rotate=self.rotation,
                    )
                except Exception as e:
                    self.output(
                        f"[{total}] Failed to write text '{text}': {e}", "ERROR"
                    )
                    continue

                try:
                    safeName: str = ""
                    for char in text:
                        if char.isalnum():
                            safeName += char
                        else:
                            safeName += "_"
                    outputFile = os.path.join(
                        self.outputPath, f"{safeName.lower()}.pdf"
                    )
                    pdf.save(outputFile)
                    successful += 1
                    self.output(f"[{total}] Wrote text '{text}' to file")
                except Exception as e:
                    self.output(
                        f"[{total}] Failed to save file for text '{text}': {e}",
                        "ERROR",
                    )
                    continue

            self.output(f"Successfully wrote {successful} out of {total} files")
